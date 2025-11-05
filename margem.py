import pandas as pd
import numpy as np
from datetime import date, datetime
import warnings
import json
import openpyxl.styles
from openpyxl.styles import Alignment

warnings.filterwarnings('ignore')

print("🚀 INICIANDO PROCESSAMENTO DE MARGEM...")

# Data fixa para o nome do arquivo
data_nome_arquivo = "testes"

REGRA_COMISSAO = {
    'geral': {
        0.00: { 
            'grupos': [
                'REDE AKKI', 'VAREJO ANDORINHA', 'VAREJO BERGAMINI', 'REDE DA PRACA', 
                'REDE DOVALE', 'REDE MERCADAO', 'REDE REIMBERG', 'REDE SEMAR', 
                'REDE TRIMAIS', 'REDE VOVO ZUZU', 'REDE BENGALA', 'VAREJO OURINHOS'
            ],
            'razoes': [
                'COMERCIO DE CARNES E ROTISSERIE DUTRA LT',
                'COMERCIO DE CARNES E ROTISSERIE DUTRA LTDA',
                'DISTRIBUIDORA E COMERCIO UAI SP LTDA',
                "GARFETO'S FORNECIMENTO DE REFEICOES LTDA", 
                "LATICINIO SOBERANO LTDA VILA ALPINA",
                "SAO LORENZO ALIMENTOS LTDA",
                "QUE DELICIA MENDES COMERCIO DE ALIMENTOS",
                "MARIANA OLIVEIRA MAZZEI",
                "LS SANTOS COMERCIO DE ALIMENTOS LTDA"
            ]
        },
        0.03: {
            'grupos': ['VAREJO CALVO', 'REDE CHAMA', 'REDE ESTRELA AZUL', 'REDE TENDA', 'REDE HIGAS']
        },
        0.01: { 
            'razoes': ['SHOPPING FARTURA VALINHOS COMERCIO LTDA']
        }
    },
    'grupos_especificos': {
        'REDE STYLLUS': {
            0.00: {
                'grupos_produto': ['TORRESMO', 'SALAME UAI', 'EMPANADOS']
            }
        },
        'REDE ROSSI': {
            0.03: {
                'codigos': [1288, 1289, 1287, 937, 1698, 1701, 1587, 1700, 1586, 1699, 943, 1735, 1624, 1134],
                'grupos_produto': ['TORRESMO']
            },
            0.01: {
                'codigos': [1265, 1266, 812, 1115, 798, 1211]
            },
            0.00: {
                'grupos_produto': [
                    'EMBUTIDOS', 'EMBUTIDOS NOBRE', 'EMBUTIDOS SADIA', 
                    'EMBUTIDOS PERDIGAO', 'EMBUTIDOS AURORA', 'EMBUTIDOS SEARA', 
                    'SALAME UAI'
                ],
                'codigos': [1139]
            },
            0.02: {
                'grupos_produto': ['MIUDOS BOVINOS', 'SUINOS', 'SALGADOS SUINOS A GRANEL'],
                'codigos': [700]
            }
        },
        'REDE PLUS': {
            0.03: {
                'grupos_produto': ['TEMPERADOS'],
                'codigos': [812]
            }
        },
        'REDE CENCOSUD': {
            0.01: {
                'grupos_produto': ['SALAME UAI']
            },
            0.03: {
                'todos_exceto': ['SALAME UAI']  # TODOS OS OUTROS PRODUTOS RECEBEM 0.03
            }
        },
        'REDE ROLDAO': {
            0.02: {
                'grupos_produto': [
                    'CONGELADOS', 'CORTES BOVINOS', 'CORTES DE FRANGO', 'EMBUTIDOS', 
                    'EMBUTIDOS AURORA', 'EMBUTIDOS NOBRE', 'EMBUTIDOS PERDIGÃO', 
                    'EMBUTIDOS SADIA', 'EMBUTIDOS SEARA', 'EMPANADOS', 
                    'KITS FEIJOADA', 'MIUDOS BOVINOS', 'SUINOS', 'TEMPERADOS'
                ]
            },
            0.00: {
                'todos_exceto': [
                    'CONGELADOS', 'CORTES BOVINOS', 'CORTES DE FRANGO', 'EMBUTIDOS', 
                    'EMBUTIDOS AURORA', 'EMBUTIDOS NOBRE', 'EMBUTIDOS PERDIGÃO', 
                    'EMBUTIDOS SADIA', 'EMBUTIDOS SEARA', 'EMPANADOS', 
                    'KITS FEIJOADA', 'MIUDOS BOVINOS', 'SUINOS', 'TEMPERADOS'
                ]
            }
        }
    },
    'razoes_especificas': {
        'PAES E DOCES LEKA LTDA': {
            0.03: [1893, 1886]
        },
        'PAES E DOCES MICHELLI LTDA': {
            0.03: [1893, 1886]
        },
        'WANDERLEY GOMES MORENO': {
            0.03: [1893, 1886]
        }
    }
}

def carregar_csv_com_codificacao(caminho, sep=';', decimal=',', thousands='.', skiprows=None):
    codificacoes = ['latin-1', 'iso-8859-1', 'cp1252', 'utf-8']
    
    for encoding in codificacoes:
        try:
            if skiprows:
                df = pd.read_csv(caminho, sep=sep, encoding=encoding, decimal=decimal, thousands=thousands, skiprows=skiprows)
            else:
                df = pd.read_csv(caminho, sep=sep, encoding=encoding, decimal=decimal, thousands=thousands)
            return df
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    
    # Última tentativa com tratamento de erros
    try:
        if skiprows:
            df = pd.read_csv(caminho, sep=sep, encoding='utf-8', decimal=decimal, thousands=thousands, 
                           skiprows=skiprows, errors='replace')
        else:
            df = pd.read_csv(caminho, sep=sep, encoding='utf-8', decimal=decimal, thousands=thousands, 
                           errors='replace')
        return df
    except Exception:
        return pd.DataFrame()

def converter_codproduto_para_int(df, coluna='CODPRODUTO'):
    df[coluna] = df[coluna].astype(str).str.strip()
    df[coluna] = df[coluna].str.replace(r'\.0$', '', regex=True)
    df[coluna] = df[coluna].str.replace(r'^0+', '', regex=True)
    df[coluna] = df[coluna].str.strip()
    
    def converter_para_int(valor):
        try:
            if pd.isna(valor) or valor == '' or valor == 'nan':
                return np.nan
            return int(float(valor))
        except (ValueError, TypeError):
            return np.nan
    
    df[coluna] = df[coluna].apply(converter_para_int)
    return df

def aplicar_regras_comissao(row):
    """
    Aplica as regras de comissão para linhas que não possuem Comissão Kg
    Retorna o valor da comissão ou None se não se aplicar
    """
    try:
        grupo = str(row['GRUPO']).strip() if pd.notna(row['GRUPO']) else ''
        razao = str(row['RAZAO']).strip() if pd.notna(row['RAZAO']) else ''
        fantasia = str(row['FANTASIA']).strip() if pd.notna(row['FANTASIA']) else ''
        grupo_produto = str(row['GRUPO PRODUTO']).strip() if pd.notna(row['GRUPO PRODUTO']) else ''
        codproduto = int(row['CODPRODUTO']) if pd.notna(row['CODPRODUTO']) else None
        
        # 1. REGRAS GERAIS
        for comissao, criterios in REGRA_COMISSAO['geral'].items():
            # Verificar grupos
            if 'grupos' in criterios and grupo in criterios['grupos']:
                return comissao
            
            # Verificar razões sociais
            if 'razoes' in criterios and (razao in criterios['razoes'] or fantasia in criterios['razoes']):
                return comissao
        
        # 2. REGRAS POR GRUPO ESPECÍFICO
        if grupo in REGRA_COMISSAO['grupos_especificos']:
            regras_grupo = REGRA_COMISSAO['grupos_especificos'][grupo]
            
            for comissao, criterios in regras_grupo.items():
                # Verificar códigos específicos
                if 'codigos' in criterios and codproduto in criterios['codigos']:
                    return comissao
                
                # Verificar grupos de produto
                if 'grupos_produto' in criterios and grupo_produto in criterios['grupos_produto']:
                    return comissao
                
                # Verificar "todos_exceto" (CENCOSUD e ROLDAO)
                if 'todos_exceto' in criterios:
                    if grupo_produto not in criterios['todos_exceto']:
                        return comissao
        
        # 3. REGRAS POR RAZÃO SOCIAL ESPECÍFICA
        for razao_especifica, regras in REGRA_COMISSAO['razoes_especificas'].items():
            if razao == razao_especifica or fantasia == razao_especifica:
                for comissao, codigos in regras.items():
                    if codproduto in codigos:
                        return comissao
        
        return None
        
    except Exception as e:
        print(f"Erro ao aplicar regras comissão: {e}")
        return None
    
def converter_data_oferta(data_str, data_referencia):
    """
    Converte datas no formato '01/ago' para datetime usando o ano da data de referência
    """
    try:
        if pd.isna(data_str):
            return pd.NaT
        
        # Se já for datetime, retorna como está
        if isinstance(data_str, (pd.Timestamp, datetime)):
            return data_str
        
        data_str = str(data_str).strip()
        
        # Mapeamento de meses abreviados
        meses_pt_br = {
            'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04', 
            'mai': '05', 'jun': '06', 'jul': '07', 'ago': '08', 
            'set': '09', 'out': '10', 'nov': '11', 'dez': '12'
        }
        
        # Tentar diferentes formatos
        formatos_tentativos = [
            '%d/%b', '%d/%B', '%d/%m', '%d/%m/%Y', 
            '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y'
        ]
        
        for formato in formatos_tentativos:
            try:
                if formato in ['%d/%b', '%d/%B']:
                    # Para formatos com mês abreviado (01/set)
                    partes = data_str.split('/')
                    if len(partes) == 2:
                        dia = partes[0].strip()
                        mes_abrev = partes[1].strip().lower()[:3]
                        
                        if mes_abrev in meses_pt_br:
                            mes_num = meses_pt_br[mes_abrev]
                            ano_atual = data_referencia.year  # Usar o ano da venda
                            data_completa = f"{dia}/{mes_num}/{ano_atual}"
                            return pd.to_datetime(data_completa, dayfirst=True, errors='coerce')
                else:
                    # Para outros formatos
                    data_convertida = pd.to_datetime(data_str, format=formato, errors='coerce')
                    if not pd.isna(data_convertida):
                        return data_convertida
            except:
                continue
        
        # Última tentativa com parser genérico
        return pd.to_datetime(data_str, dayfirst=True, errors='coerce')
        
    except Exception as e:
        print(f"Erro ao converter data oferta '{data_str}': {e}")
        return pd.NaT

def buscar_oferta_vog(row, ofertas_off, ofertas_cb):
    """
    Busca ofertas conforme a regra:
    - CORTES BOVINOS: sempre procura em OFF_VOG_CB
    - Todo resto: sempre procura em OFF_VOG
    """
    try:
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else ''
        data_venda = row['DATA']
        preco_venda = row['Preço Venda']
        grupo_produto = str(row['GRUPO PRODUTO']).strip() if pd.notna(row['GRUPO PRODUTO']) else ''
        
        if not codproduto or codproduto == 'nan' or pd.isna(data_venda) or pd.isna(preco_venda) or preco_venda == 0:
            return None
        
        # Converter data da venda para Timestamp se necessário
        if isinstance(data_venda, date):
            data_venda = pd.Timestamp(data_venda)
        
        try:
            codproduto_int = int(float(codproduto))
        except (ValueError, TypeError):
            return None
        
        # REGRA: CORTES BOVINOS sempre procura em OFF_VOG_CB
        if grupo_produto == "CORTES BOVINOS":
            comissao = buscar_oferta_cb(row, ofertas_cb, codproduto_int, data_venda, preco_venda)
            # Se não encontrou oferta específica em CB, aplica 2% padrão para CORTES BOVINOS
            if comissao is None:
                return 0.02
            return comissao
        else:
            # TODO O RESTO sempre procura em OFF_VOG
            comissao = buscar_oferta_off(row, ofertas_off, codproduto_int, data_venda, preco_venda, grupo_produto)
            # Se não encontrou oferta específica em OFF_VOG, aplica 3% padrão
            if comissao is None:
                return 0.03
            return comissao
        
    except Exception as e:
        print(f"❌ ERRO ao buscar oferta para produto {codproduto}: {e}")
        return None
    
def buscar_oferta_off(row, ofertas_off, codproduto_int, data_venda, preco_venda, grupo_produto):
    """Busca ofertas na aba OFF_VOG para produtos que NÃO são CORTES BOVINOS"""
    if ofertas_off.empty or 'COD' not in ofertas_off.columns or 'DT_REF_OFF' not in ofertas_off.columns:
        print(f"❌ OFF_VOG não disponível para produto {codproduto_int}")
        return None  # Retorna None para aplicar 3% padrão
    
    ofertas_cod = ofertas_off[ofertas_off['COD'] == codproduto_int].copy()
    
    if ofertas_cod.empty:
        print(f"❌ Produto {codproduto_int} não encontrado em OFF_VOG")
        return None  # Retorna None para aplicar 3% padrão
    
    ofertas_cod = ofertas_cod.copy()
    ofertas_cod['DT_REF_OFF_CONVERTED'] = ofertas_cod['DT_REF_OFF'].apply(
        lambda x: converter_data_oferta(x, data_venda)
    )
    
    ofertas_cod = ofertas_cod.dropna(subset=['DT_REF_OFF_CONVERTED'])
    
    if ofertas_cod.empty:
        print(f"❌ Nenhuma data válida para produto {codproduto_int} em OFF_VOG")
        return None
    
    ofertas_cod = ofertas_cod.sort_values('DT_REF_OFF_CONVERTED', ascending=False)
    
    # Buscar data exata ou anterior mais próxima
    ofertas_validas = ofertas_cod[ofertas_cod['DT_REF_OFF_CONVERTED'] <= data_venda]
    
    if not ofertas_validas.empty:
        oferta_mais_recente = ofertas_validas.iloc[0]
        
        # Buscar coluna 3%
        coluna_3pct = None
        for col in oferta_mais_recente.index:
            if '3%' in str(col) or '3.00%' in str(col):
                coluna_3pct = col
                break
        
        if coluna_3pct and pd.notna(oferta_mais_recente[coluna_3pct]) and oferta_mais_recente[coluna_3pct] != 0:
            preco_oferta_3pct = oferta_mais_recente[coluna_3pct]
            
            # Lógica VOG: 3% se >=, 1% se <
            if preco_venda >= preco_oferta_3pct:
                return 0.03
            else:
                return 0.01
    
    # Se não encontrou oferta válida em OFF_VOG
    print(f"❌ Nenhuma oferta válida em OFF_VOG para produto {codproduto_int}")
    return None

def buscar_oferta_cb(row, ofertas_cb, codproduto_int, data_venda, preco_venda):
    """Busca ofertas na aba OFF_VOG_CB para produtos CORTES BOVINOS"""
    if ofertas_cb.empty or 'CD_PROD' not in ofertas_cb.columns or 'DT_REF' not in ofertas_cb.columns:
        print(f"❌ OFF_VOG_CB não disponível para CORTES BOVINOS {codproduto_int}")
        return None  # Retorna None para aplicar 2% padrão
    
    ofertas_cod = ofertas_cb[ofertas_cb['CD_PROD'] == codproduto_int].copy()
    
    if ofertas_cod.empty:
        print(f"❌ CORTES BOVINOS {codproduto_int} não encontrado em OFF_VOG_CB")
        return None  # Retorna None para aplicar 2% padrão
    
    ofertas_cod = ofertas_cod.copy()
    ofertas_cod['DT_REF_CONVERTED'] = ofertas_cod['DT_REF'].apply(
        lambda x: converter_data_oferta(x, data_venda)
    )
    
    ofertas_cod = ofertas_cod.dropna(subset=['DT_REF_CONVERTED'])
    
    if ofertas_cod.empty:
        print(f"❌ Nenhuma data válida para CORTES BOVINOS {codproduto_int} em OFF_VOG_CB")
        return None
    
    ofertas_cod = ofertas_cod.sort_values('DT_REF_CONVERTED', ascending=False)
    
    # Buscar data exata ou anterior mais próxima
    ofertas_validas = ofertas_cod[ofertas_cod['DT_REF_CONVERTED'] <= data_venda]
    
    if not ofertas_validas.empty:
        oferta_mais_recente = ofertas_validas.iloc[0]
        
        # Buscar coluna 2% - em OFF_VOG_CB usamos a coluna "2%"
        coluna_2pct = '2%' if '2%' in oferta_mais_recente.index else None
        
        if coluna_2pct and pd.notna(oferta_mais_recente[coluna_2pct]) and oferta_mais_recente[coluna_2pct] != 0:
            preco_oferta_2pct = oferta_mais_recente[coluna_2pct]
            
            # Lógica CB: 2% se >=, 1% se <
            if preco_venda >= preco_oferta_2pct:
                return 0.02
            else:
                return 0.01
        else:
            # Se não encontrou coluna 2%, aplicar 2% padrão para CORTES BOVINOS
            return 0.02
    
    # Se não encontrou oferta válida em OFF_VOG_CB
    print(f"❌ Nenhuma oferta válida em OFF_VOG_CB para CORTES BOVINOS {codproduto_int}")
    return None
    
# Carregar fechamento
fechamento = carregar_csv_com_codificacao(r"C:\Users\win11\Downloads\fechamento_processado.csv")

# Carregar cancelados (com skiprows=2)
cancelados = carregar_csv_com_codificacao(r"S:\hor\arquivos\gustavo\can.csv", skiprows=2)

# Carregar devoluções (arquivo onde vamos buscar o PESO para QTDE AJUSTADA)
devolucoes = carregar_csv_com_codificacao(r"S:\hor\excel\20251101.csv")

# Carregar custos_produtos (Excel)
try:
    custos_produtos = pd.read_excel(r"C:\Users\win11\Downloads\Custos de produtos - Novembro.xlsx", sheet_name='Base', dtype=str)
except Exception:
    custos_produtos = pd.DataFrame()

# Carregar LOURENCINI
try:
    lourencini = pd.read_excel(r"C:\Users\win11\Downloads\LOURENCINI.xlsx")
    required_cols = ['COD', '0,15', '0,3', '0,5', '0,7', '1', 'Data', 'Data_fim']
    if all(col in lourencini.columns for col in required_cols):
        lourencini['COD'] = lourencini['COD'].astype(str).str.strip()
        lourencini['COD'] = lourencini['COD'].str.replace(r'\.0$', '', regex=True)
        lourencini['COD'] = lourencini['COD'].str.replace(r'^0+', '', regex=True)
        lourencini['COD'] = lourencini['COD'].str.strip()
        lourencini = lourencini[lourencini['COD'] != '']
        lourencini = lourencini[lourencini['COD'] != 'nan']
        lourencini = lourencini.dropna(subset=['COD'])
        
        def converter_para_int_se_possivel(valor):
            try:
                if pd.isna(valor) or valor == '':
                    return np.nan
                return int(float(valor))
            except (ValueError, TypeError):
                return np.nan
        
        lourencini['COD'] = lourencini['COD'].apply(converter_para_int_se_possivel)
        lourencini = lourencini.dropna(subset=['COD'])
        
        colunas_preco = ['0,15', '0,3', '0,5', '0,7', '1']
        for col in colunas_preco:
            lourencini[col] = pd.to_numeric(lourencini[col], errors='coerce')
        
        lourencini['Data'] = pd.to_datetime(lourencini['Data'], errors='coerce', dayfirst=True)
        if 'Data_fim' in lourencini.columns:
            lourencini['Data_fim'] = pd.to_datetime(lourencini['Data_fim'], errors='coerce', dayfirst=True)
        
        lourencini = lourencini.sort_values('Data')
    else:
        lourencini = pd.DataFrame()
except Exception:
    lourencini = pd.DataFrame()
    print("⚠️ Erro ao carregar LOURENCINI")

# Carregar OFERTAS_VOG 
try:
    ofertas_off = pd.read_excel(r"C:\Users\win11\Downloads\OFERTAS_VOG.xlsx", sheet_name='OFF_VOG')
except Exception:
    ofertas_off = pd.DataFrame()
    print("⚠️ Erro ao carregar OFF_VOG")

try:
    ofertas_cb = pd.read_excel(r"C:\Users\win11\Downloads\OFERTAS_VOG.xlsx", sheet_name='OFF_VOG_CB')
except Exception:
    ofertas_cb = pd.DataFrame()
    print("⚠️ Erro ao carregar OFF_VOG_CB")

# Verificar se os DataFrames essenciais foram carregados
if fechamento.empty:
    print("❌ CRÍTICO: DataFrame 'fechamento' está vazio!")
    exit()

if cancelados.empty:
    print("⚠️ AVISO: DataFrame 'cancelados' está vazio!")

if devolucoes.empty:
    print("⚠️ AVISO: DataFrame 'devolucoes' está vazio!")

print("✅ TODOS OS ARQUIVOS CARREGADOS")
print(f"📊 Tamanhos: fechamento={len(fechamento)}, cancelados={len(cancelados)}, devoluções={len(devolucoes)}")

if not devolucoes.empty:
    devolucoes['HISTORICO'] = devolucoes['HISTORICO'].astype(str).str.strip()
    devolucoes_filtradas = devolucoes[
        (devolucoes['HISTORICO'] == '51') | 
        (devolucoes['HISTORICO'] == '68')
    ].copy()
    devolucoes = devolucoes_filtradas
else:
    print("⚠️ AVISO: DataFrame 'devolucoes' está vazio, não foi possível filtrar por HISTORICO")

# Renomear colunas e processar custos_produtos
rename_mapping = {
    'PRODUTO': 'CODPRODUTO', 'DATA': 'DATA', 'PCS': 'QTDE', 'KGS': 'PESO_KGS', 
    'CUSTO': 'CUSTO', 'FRETE': 'FRETE', 'PRODUÇÃO': 'PRODUÇÃO', 'TOTAL': 'TOTAL', 
    'QTD': 'QTD', 'PESO': 'PESO'
}
custos_produtos = custos_produtos.rename(columns=rename_mapping)

colunas_numericas = ['PESO_KGS', 'CUSTO', 'FRETE', 'PRODUÇÃO', 'TOTAL', 'QTD', 'PESO']
for coluna in colunas_numericas:
    if coluna in custos_produtos.columns:
        try:
            if custos_produtos[coluna].dtype != 'object':
                custos_produtos[coluna] = custos_produtos[coluna].astype(str)
            custos_produtos[coluna] = custos_produtos[coluna].apply(
                lambda x: str(x).replace(',', '.').replace(' ', '') if pd.notna(x) else x
            )
            custos_produtos[coluna] = pd.to_numeric(custos_produtos[coluna], errors='coerce')
        except Exception:
            pass

custos_produtos['DATA'] = pd.to_datetime(custos_produtos['DATA'], errors='coerce', dayfirst=True)
custos_produtos = converter_codproduto_para_int(custos_produtos)

# Processar dados
notas_canceladas = cancelados['NUMERO'].tolist() if not cancelados.empty else []

if not devolucoes.empty:
    devolucoes_filtro = devolucoes[
        (devolucoes['DESCRICAO'] == "DEV VENDA C/ FIN S/ EST") | 
        (devolucoes['HISTORICO'] == "68")
    ]
    devolucoes_var = devolucoes_filtro[['ROMANEIO', 'NOTA FISCAL']].values.tolist()

    vendas_filtro = devolucoes[
        (devolucoes['DESCRICAO'] == "VENDA") | 
        (devolucoes['HISTORICO'] == "51")
    ]
    vendas_var = vendas_filtro[['ROMANEIO', 'NOTA FISCAL']].values.tolist()
else:
    devolucoes_var = []
    vendas_var = []


fechamento_sem_cancelados = fechamento[~fechamento['NF-E'].isin(notas_canceladas)].copy()

if not devolucoes.empty:
    combinacoes_validas = devolucoes[['ROMANEIO', 'NOTA FISCAL']].drop_duplicates()
    
    fechamento_sem_cancelados['CHAVE_ROMANEIO_NF'] = (
        fechamento_sem_cancelados['ROMANEIO'].astype(str) + "_" + 
        fechamento_sem_cancelados['NF-E'].astype(str)
    )
    
    combinacoes_validas['CHAVE_ROMANEIO_NF'] = (
        combinacoes_validas['ROMANEIO'].astype(str) + "_" + 
        combinacoes_validas['NOTA FISCAL'].astype(str)
    )
    
    chaves_validas = set(combinacoes_validas['CHAVE_ROMANEIO_NF'].tolist())
    
    fechamento_filtrado = fechamento_sem_cancelados[
        fechamento_sem_cancelados['CHAVE_ROMANEIO_NF'].isin(chaves_validas)
    ].copy()

    fechamento_filtrado = fechamento_filtrado.drop('CHAVE_ROMANEIO_NF', axis=1)
    fechamento_sem_cancelados = fechamento_filtrado
else:
    print("⚠️ AVISO: Não foi possível filtrar por histórico - devoluções vazio")
    fechamento_sem_cancelados = fechamento_sem_cancelados

# Criar dicionário de custos
custos_dict = {}
custos_data = custos_produtos.to_dict('records')

for i, row in enumerate(custos_data):
    try:
        data_value = row.get('DATA', None)
        codproduto_value = row.get('CODPRODUTO', None)
        
        if data_value is None or pd.isna(data_value) or (isinstance(data_value, str) and data_value.strip() == ''):
            continue
        if codproduto_value is None or pd.isna(codproduto_value) or (isinstance(codproduto_value, str) and codproduto_value.strip() == ''):
            continue
        
        codproduto = str(codproduto_value).strip()
        data_key = pd.to_datetime(data_value, errors='coerce', dayfirst=True)
        if pd.isna(data_key):
            continue
        data_key = data_key.date()
        
        custo_val = float(row.get('CUSTO', 0)) if pd.notna(row.get('CUSTO')) and str(row.get('CUSTO', '')).strip() != '' else 0
        peso_val = float(row.get('PESO', 1)) if pd.notna(row.get('PESO')) and str(row.get('PESO', '')).strip() != '' else 1
        producao_val = float(row.get('PRODUÇÃO', 0)) if pd.notna(row.get('PRODUÇÃO')) and str(row.get('PRODUÇÃO', '')).strip() != '' else 0
        frete_val = float(row.get('FRETE', 0)) if pd.notna(row.get('FRETE')) and str(row.get('FRETE', '')).strip() != '' else 0
        qtd_val = float(row.get('QTD', 0)) if pd.notna(row.get('QTD')) and str(row.get('QTD', '')).strip() != '' else 0
        
        custos_dict[(codproduto, data_key)] = {
            'QTD': qtd_val, 'PESO': peso_val, 'CUSTO': custo_val, 
            'FRETE': frete_val, 'PRODUÇÃO': producao_val
        }
    except Exception:
        continue

# Dicionários para lookup
quinzena_dict = {}
fechamento['PK'] = fechamento['ROMANEIO'].astype(str) + "_" + fechamento['NF-E'].astype(str) + "_" + fechamento['CODPRODUTO'].astype(str)
for _, row in fechamento.iterrows():
    try:
        if pd.notna(row['QUINZENA']):
            quinzena_value = str(row['QUINZENA'])
            if quinzena_value == "Primeira Quinzena":
                quinzena_dict[row['PK']] = "1ª Quinzena"
            elif quinzena_value == "Segunda Quinzena":
                quinzena_dict[row['PK']] = "2ª Quinzena"
            else:
                quinzena_dict[row['PK']] = quinzena_value
    except:
        continue

comissao_regra_dict = {}
for _, row in fechamento.iterrows():
    try:
        key = (int(row['ROMANEIO']), int(row['NF-E']), int(row['CODPRODUTO']))
        comissao_regra_dict[key] = row['P.COM'] if pd.notna(row['P.COM']) else 0
    except:
        continue

fechamento_pk_dict = {}
for _, row in fechamento.iterrows():
    pk = str(row['ROMANEIO']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
    desconto_verificado = row['Desconto verificado'] if 'Desconto verificado' in fechamento.columns and pd.notna(row['Desconto verificado']) else np.nan
    
    quinzena_value = row['QUINZENA'] if 'QUINZENA' in fechamento.columns and pd.notna(row['QUINZENA']) else ""
    if quinzena_value == "Primeira Quinzena":
        quinzena_value = "1ª Quinzena"
    elif quinzena_value == "Segunda Quinzena":
        quinzena_value = "2ª Quinzena"
    
    fechamento_pk_dict[pk] = {
        'ESCRITORIO': row['ESCRITORIO'] if 'ESCRITORIO' in fechamento.columns and pd.notna(row['ESCRITORIO']) else np.nan,
        'VLR ICMS': row['VLR ICMS'] if 'VLR ICMS' in fechamento.columns and pd.notna(row['VLR ICMS']) else np.nan,
        'PRECO VENDA': row['PRECO VENDA'] if 'PRECO VENDA' in fechamento.columns and pd.notna(row['PRECO VENDA']) else np.nan,
        'QUINZENA': quinzena_value,
        'DESCONTO_VERIFICADO': desconto_verificado,
        'MOV': row['Mov'] if 'Mov' in fechamento.columns and pd.notna(row['Mov']) else "",
        'MOV_V2': row['Mov V2'] if 'Mov V2' in fechamento.columns and pd.notna(row['Mov V2']) else ""
    }

fechamento_nf_dict = {}
for _, row in fechamento.iterrows():
    if pd.notna(row['NF-E']):
        fechamento_nf_dict[int(row['NF-E'])] = row['Mov'] if 'Mov' in fechamento.columns and pd.notna(row['Mov']) else ""

# CRIAR DICIONÁRIO PARA QTDE AJUSTADA A PARTIR DO ARQUIVO DE DEVOLUÇÕES
qtde_ajustada_dict = {}

if not devolucoes.empty:
    colunas_necessarias = ['NOTA FISCAL', 'ROMANEIO', 'PRODUTO', 'PESO']
    colunas_existentes = [col for col in colunas_necessarias if col in devolucoes.columns]
    
    if len(colunas_existentes) == 4:
        for _, row in devolucoes.iterrows():
            try:
                nota_fiscal = row['NOTA FISCAL']
                romaneio = row['ROMANEIO']
                produto = row['PRODUTO']
                peso = row['PESO']
                
                if pd.notna(nota_fiscal) and pd.notna(romaneio) and pd.notna(produto) and pd.notna(peso):
                    nota_fiscal_str = str(nota_fiscal).strip()
                    romaneio_str = str(romaneio).strip()
                    produto_str = str(produto).strip()
                    
                    chave = (nota_fiscal_str, romaneio_str, produto_str)
                    
                    try:
                        peso_float = float(str(peso).replace(',', '.'))
                        qtde_ajustada_dict[chave] = peso_float
                    except (ValueError, TypeError):
                        continue
                        
            except Exception:
                continue
    else:
        qtde_ajustada_dict = {}
else:
    qtde_ajustada_dict = {}

# Criar base_df
base_df = pd.DataFrame()

# Preencher colunas básicas
base_df['CF'] = fechamento_sem_cancelados.apply(
    lambda row: 'DEV' if any([str(row['ROMANEIO']) == str(dev[0]) and str(row['NF-E']) == str(dev[1]) for dev in devolucoes_var]) 
    else row['LOJA'], axis=1
)
base_df['RAZAO'] = fechamento_sem_cancelados['RAZAO']
base_df['FANTASIA'] = fechamento_sem_cancelados['FANTASIA']
base_df['GRUPO'] = fechamento_sem_cancelados['GRUPO']
base_df['OS'] = fechamento_sem_cancelados['ROMANEIO']
base_df['NF-E'] = fechamento_sem_cancelados['NF-E']
base_df['CF_NF'] = fechamento_sem_cancelados['CF_NF'].fillna("")
base_df['DATA'] = pd.to_datetime(fechamento_sem_cancelados['DATA'], dayfirst=True, errors='coerce').dt.date
base_df['VENDEDOR'] = fechamento_sem_cancelados['VENDEDOR']

# Converter CODPRODUTO para inteiro
base_df['CODPRODUTO'] = fechamento_sem_cancelados['CODPRODUTO'].astype(str)
base_df = converter_codproduto_para_int(base_df)

base_df['GRUPO PRODUTO'] = fechamento_sem_cancelados['GRUPO PRODUTO']
base_df['DESCRICAO'] = fechamento_sem_cancelados['DESCRICAO']
base_df['QTDE'] = fechamento_sem_cancelados['QTDE']  
base_df['QTDE REAL'] = fechamento_sem_cancelados['QTDE REAL'] 
base_df['CUSTO EM SISTEMA'] = fechamento_sem_cancelados['CUSTO']
base_df['Val Pis'] = fechamento_sem_cancelados['VLR PIS'].fillna(0) if 'VLR PIS' in fechamento_sem_cancelados.columns else 0
base_df['VLRCOFINS'] = fechamento_sem_cancelados['VLR COFINS'].fillna(0) if 'VLR COFINS' in fechamento_sem_cancelados.columns else 0
base_df['IRPJ'] = fechamento_sem_cancelados['IRPJ'].fillna(0) if 'IRPJ' in fechamento_sem_cancelados.columns else 0
base_df['CSLL'] = fechamento_sem_cancelados['CSLL'].fillna(0) if 'CSLL' in fechamento_sem_cancelados.columns else 0
base_df['VL ICMS'] = fechamento_sem_cancelados['VLR ICMS'] if 'VLR ICMS' in fechamento_sem_cancelados.columns else 0
base_df['Preço Venda'] = fechamento_sem_cancelados['PRECO VENDA'] if 'PRECO VENDA' in fechamento_sem_cancelados.columns else 0

base_df['PK'] = base_df['OS'].astype(str) + "_" + base_df['NF-E'].astype(str) + "_" + base_df['CODPRODUTO'].astype(str)
base_df['Quinzena'] = base_df['PK'].map(lambda x: quinzena_dict.get(x, ""))
base_df['GRUPO'] = base_df['GRUPO'].fillna('VAREJO')  

# FUNÇÃO MODIFICADA PARA CALCULAR QTDE AJUSTADA (COM CONDIÇÃO PARA VALORES NEGATIVOS)
def calcular_qtde_ajustada(row):
    try:
        # PRIMEIRO: Verificar se é CF = "ESP" para aplicar lógica especial
        cf_valor = str(row['CF']).strip() if pd.notna(row['CF']) else ""
        
        # SE FOR ESP: Aplicar a lógica atual de buscar no dicionário e custos
        if cf_valor == "ESP":
            nf_e = str(row['NF-E']).strip() if pd.notna(row['NF-E']) else ""
            os_val = str(row['OS']).strip() if pd.notna(row['OS']) else ""
            codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else ""
            
            chave = (nf_e, os_val, codproduto)
            
            if chave in qtde_ajustada_dict:
                peso_encontrado = qtde_ajustada_dict[chave]
                return peso_encontrado
            
            if row['QTDE REAL'] <= 0:
                return row['QTDE REAL']
            
            data = row['DATA']
            
            if codproduto is None or data is None:
                return row['QTDE REAL']
                
            custo_info = custos_dict.get((codproduto, data), {})
            qtd = custo_info.get('QTD', 1)
            
            if qtd > 1:
                resultado = row['QTDE'] * qtd
            else:
                resultado = row['QTDE REAL']
                
            return resultado
        
        # SE FOR DEV: Manter o mesmo valor de QTDE REAL (não aplicar lógica especial)
        elif cf_valor == "DEV":
            return row['QTDE REAL']
            
        # PARA OUTROS CASOS: Aplicar lógica padrão
        else:
            nf_e = str(row['NF-E']).strip() if pd.notna(row['NF-E']) else ""
            os_val = str(row['OS']).strip() if pd.notna(row['OS']) else ""
            codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else ""
            
            chave = (nf_e, os_val, codproduto)
            
            if chave in qtde_ajustada_dict:
                peso_encontrado = qtde_ajustada_dict[chave]
                
                # Apenas para DEV (não ESP) aplicar negativo
                if cf_valor == "DEV":
                    peso_encontrado = -abs(peso_encontrado)
                
                return peso_encontrado
            
            if row['QTDE REAL'] <= 0:
                return row['QTDE REAL']
            
            data = row['DATA']
            
            if codproduto is None or data is None:
                return row['QTDE REAL']
                
            custo_info = custos_dict.get((codproduto, data), {})
            qtd = custo_info.get('QTD', 1)
            
            if qtd > 1:
                resultado = row['QTDE'] * qtd
            else:
                resultado = row['QTDE REAL']
                
            # Apenas para DEV (não ESP) aplicar negativo
            if cf_valor == "DEV" and resultado > 0:
                resultado = -resultado
                
            return resultado
            
    except Exception:
        return row['QTDE REAL']

def calcular_qtde_real2(row):
    try:
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else None
        data = row['DATA']
        
        if codproduto is None or data is None:
            return np.nan
            
        custo_info = custos_dict.get((codproduto, data), {})
        peso = custo_info.get('PESO', 1)
        
        if row['QTDE REAL'] < 0:
            return -row['QTDE AJUSTADA'] * peso
        else:
            return row['QTDE AJUSTADA'] * peso
    except:
        return np.nan

def buscar_custo(row):
    try:
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else None
        data = row['DATA']
        
        if codproduto is None or data is None:
            return np.nan
            
        key = (codproduto, data)
        if key in custos_dict:
            custo = custos_dict[key].get('CUSTO', 0)
            return custo if custo != 0 else np.nan
        else:
            return np.nan
    except:
        return np.nan

def buscar_frete(row):
    if row['FANTASIA'] in ["PASSOS ALIMENTOS LTDA", "AGELLE ARMAZEM E LOGISTICA LTDA", 
                           "GEMEOS REORESENTACOES", "REAL DISTRIBUIDORA"]:
        return 0
        
    try:
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else None
        data = row['DATA']
        
        if codproduto is None or data is None:
            return 0
            
        custo_info = custos_dict.get((codproduto, data), {})
        return custo_info.get('FRETE', 0)
    except:
        return 0

def buscar_producao(row):
    try:
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else None
        data = row['DATA']
        
        if codproduto is None or data is None:
            return 0
            
        custo_info = custos_dict.get((codproduto, data), {})
        return custo_info.get('PRODUÇÃO', 0)
    except:
        return 0

# Aplicar funções
base_df['QTDE AJUSTADA'] = base_df.apply(calcular_qtde_ajustada, axis=1)
base_df['QTDE REAL2'] = base_df.apply(calcular_qtde_real2, axis=1)
base_df['CUSTO'] = base_df.apply(buscar_custo, axis=1)
base_df['Custo total'] = base_df['CUSTO'] * base_df['QTDE AJUSTADA']
base_df['Frete'] = base_df.apply(buscar_frete, axis=1)
base_df['Produção'] = base_df.apply(buscar_producao, axis=1)

# Escritório
if 'ESCRITORIO' in fechamento_sem_cancelados.columns:
    base_df['Escritório'] = fechamento_sem_cancelados['ESCRITORIO'].fillna(0) / 100
else:
    base_df['Escritório'] = 0

# MODIFICAÇÃO: Transformar 4% e 0% em 3.5%
base_df['Escritório'] = base_df['Escritório'].apply(
    lambda x: 0.035 if abs(x - 0.04) < 0.001 or abs(x - 0.00) < 0.001 else x
)

# Desconto - CORREÇÃO COMPLETA
base_df['Desc Perc'] = 0

if 'DESCONTO' in fechamento_sem_cancelados.columns:
    # PRIMEIRO: Preencher valores vazios/NaN com 0 na coluna DESCONTO
    fechamento_sem_cancelados['DESCONTO'] = fechamento_sem_cancelados['DESCONTO'].fillna(0)
    
    # Converter para string e tratar valores vazios
    fechamento_sem_cancelados['DESCONTO'] = fechamento_sem_cancelados['DESCONTO'].apply(
        lambda x: '0' if pd.isna(x) or str(x).strip() == '' else str(x)
    )
    
    # Criar chave única para merge
    fechamento_sem_cancelados['MERGE_KEY'] = (
        fechamento_sem_cancelados['ROMANEIO'].astype(str) + "_" + 
        fechamento_sem_cancelados['NF-E'].astype(str) + "_" + 
        fechamento_sem_cancelados['CODPRODUTO'].astype(str)
    )
    
    base_df['MERGE_KEY'] = (
        base_df['OS'].astype(str) + "_" + 
        base_df['NF-E'].astype(str) + "_" + 
        base_df['CODPRODUTO'].astype(str)
    )
    
    # Criar dicionário de descontos
    desconto_dict = {}
    for _, row in fechamento_sem_cancelados.iterrows():
        key = row['MERGE_KEY']
        desconto_val = row['DESCONTO']
        
        try:
            # Já garantimos que não tem valores vazios, mas vamos limpar
            desconto_str = str(desconto_val).strip().replace(',', '.')
            # Remover possíveis caracteres não numéricos (exceto ponto e sinal negativo)
            desconto_str = ''.join(c for c in desconto_str if c.isdigit() or c in '.-')
            
            if not desconto_str or desconto_str == '.':
                desconto_float = 0
            else:
                desconto_float = float(desconto_str)
                # Se o valor for maior que 1, provavelmente está em porcentagem (5 → 5%)
                if desconto_float > 1:
                    desconto_float = desconto_float / 100
            
            desconto_dict[key] = desconto_float
            
        except (ValueError, TypeError) as e:
            print(f"⚠️ Erro ao converter desconto '{desconto_val}': {e}")
            desconto_dict[key] = 0
    
    # Aplicar descontos
    base_df['Desc Perc'] = base_df['MERGE_KEY'].map(desconto_dict).fillna(0)
    
    # Remover coluna temporária
    base_df = base_df.drop('MERGE_KEY', axis=1)

# VERIFICAÇÃO - mostrar alguns exemplos de descontos
print("🔍 AMOSTRA DE DESCONTOS APLICADOS:")
amostra_descontos = base_df[['OS', 'NF-E', 'CODPRODUTO', 'Desc Perc']].head(15)
print(amostra_descontos)

# Garantir que Desc. Valor seja calculado corretamente
base_df['Desc. Valor'] = base_df.apply(
    lambda row: 0 if (row['CF'] == "DEV" or row['GRUPO'] == "TENDA") 
    else row['QTDE AJUSTADA'] * row['Preço Venda'] * row['Desc Perc'], 
    axis=1
)

# Fat. Bruto
base_df['Fat. Bruto'] = base_df.apply(
    lambda row: -row['QTDE AJUSTADA'] * row['Preço Venda'] if row['CF'] == "DEV"
    else row['QTDE AJUSTADA'] * row['Preço Venda'], axis=1
)

# Aliq Icms
base_df['Aliq Icms'] = base_df.apply(
    lambda row: round(row['VL ICMS'] / row['Fat. Bruto'], 2) if (row['Fat. Bruto'] != 0 and pd.notna(row['VL ICMS'])) 
    else 0, axis=1
)
base_df['Aliq Icms'] = base_df['Aliq Icms'].replace([np.inf, -np.inf], 0)

# Custo real
base_df['Custo real'] = base_df.apply(
    lambda row: 0 if (pd.isna(row['QTDE AJUSTADA']) or row['QTDE AJUSTADA'] <= 0 or 
                     pd.isna(row['CUSTO']) or pd.isna(row['Aliq Icms']))
    else row['CUSTO'] - (row['CUSTO'] * row['Aliq Icms']), axis=1
)

# Fat Liquido
base_df['Fat Liquido'] = base_df.apply(
    lambda row: row['QTDE AJUSTADA'] * (row['Preço Venda'] - row['Preço Venda'] * row['Desc Perc']) if row['CF'] != "DEV"
    else -row['QTDE AJUSTADA'] * (row['Preço Venda'] - row['Preço Venda'] * row['Desc Perc']), axis=1
)

# Aniversário
base_df['Aniversário'] = base_df.apply(
    lambda row: 0 if row['CF'] == "DEV" else row['Fat. Bruto'] * 0.01, axis=1
)

def calcular_comissao_kg_simplificada(row):
    """Calcula Comissão Kg de forma simplificada"""
    try:
        if row['CF'] == "DEV":
            return 0
        
        vendedor = str(row['VENDEDOR']).strip() if pd.notna(row['VENDEDOR']) else ''
        codproduto = str(row['CODPRODUTO']).strip() if pd.notna(row['CODPRODUTO']) else ''
        grupo = str(row['GRUPO']).strip() if pd.notna(row['GRUPO']) else ''
        
        # REGRAS ESPECÍFICAS POR VENDEDOR (COMISSÃO POR KG)
        regras_vendedores = {
            "LUIZ FERNANDO VOLTERO BARBOSA": {
                "812": {"REDE CHAMA": 3, "REDE PARANA": 3, "REDE PLUS": 2}
            },
            "FELIPE RAMALHO GOMES": {
                "700": {"REDE PEDREIRA": 2, "VAREJO BERGAMINI": 0.5}
            },
            "VALDENIR VOLTERO": {
                "812": {"REDE RICOY": 1}, 
                "937": {"REDE RICOY": 0.5}, 
                "1624": {"REDE RICOY": 0.5}
            },
            "ROSE VOLTERO": {"812": 2},
            "VERA LUCIA MUNIZ": {"812": 2},
            "PAMELA FERREIRA VIEIRA": {"812": 2}
        }
        
        if vendedor in regras_vendedores:
            vendedor_regras = regras_vendedores[vendedor]
            if codproduto in vendedor_regras:
                regra = vendedor_regras[codproduto]
                if isinstance(regra, dict):
                    comissao_especifica = regra.get(grupo, None)
                    if comissao_especifica is not None:
                        return comissao_especifica
                else:
                    return regra
        
        # REGRA LOURENCINI (COMISSÃO POR KG)
        if row['GRUPO'] == "REDE LOURENCINI" and not lourencini.empty:
            data_venda = row['DATA']
            preco_venda = row['Preço Venda']
            
            if not codproduto or codproduto == 'nan' or pd.isna(data_venda) or pd.isna(preco_venda) or preco_venda == 0:
                return 0
            
            try:
                codproduto_int = int(codproduto)
            except (ValueError, TypeError):
                return 0
            
            lourencini_filtrado = lourencini[lourencini['COD'] == codproduto_int]
            if lourencini_filtrado.empty:
                return 0
            
            data_venda_dt = pd.Timestamp(data_venda)
            lourencini_row = None
            
            if 'Data_fim' in lourencini_filtrado.columns:
                lourencini_periodo = lourencini_filtrado[
                    (lourencini_filtrado['Data'] <= data_venda_dt) & 
                    (lourencini_filtrado['Data_fim'] >= data_venda_dt)
                ]
                if not lourencini_periodo.empty:
                    lourencini_row = lourencini_periodo.iloc[0]
            
            if lourencini_row is None:
                lourencini_anteriores = lourencini_filtrado[lourencini_filtrado['Data'] <= data_venda_dt]
                if not lourencini_anteriores.empty:
                    lourencini_anteriores = lourencini_anteriores.sort_values('Data', ascending=False)
                    lourencini_row = lourencini_anteriores.iloc[0]
                else:
                    lourencini_filtrado = lourencini_filtrado.sort_values('Data', ascending=True)
                    lourencini_row = lourencini_filtrado.iloc[0]
            
            colunas_comissao = ['0,15', '0,3', '0,5', '0,7', '1']
            melhor_comissao = None
            menor_diferenca = float('inf')
            
            for coluna in colunas_comissao:
                valor_na_tabela = lourencini_row[coluna]
                
                if pd.notna(valor_na_tabela) and valor_na_tabela != 0:
                    diferenca = abs(preco_venda - valor_na_tabela)
                    
                    if diferenca < menor_diferenca:
                        menor_diferenca = diferenca
                        melhor_comissao = float(coluna.replace(',', '.'))
            
            if melhor_comissao is not None:
                return melhor_comissao
        
        return 0
        
    except Exception as e:
        print(f"Erro comissão kg: {e}")
        return 0

def calcular_p_com_final(row):
    """
    Calcula P.Com na seguinte ordem:
    1. Se tem Comissão Kg: P.Com = Comissão Kg / Preço Venda
    2. Se não, aplica regras de comissão
    3. Se nenhuma regra se aplica: busca ofertas
    4. Se nada encontrado: 3% padrão (ou -3% para DEV)
    """
    try:
        is_devolucao = str(row['CF']).startswith('DEV')
        
        # 1. PRIMEIRO: COMISSÃO POR KG
        comissao_kg = calcular_comissao_kg_simplificada(row)
        
        if comissao_kg is not None and comissao_kg > 0 and row['Preço Venda'] > 0:
            p_com_calculado = comissao_kg / row['Preço Venda']
            return -p_com_calculado if is_devolucao else p_com_calculado
        
        # 2. SEGUNDO: REGRAS DE COMISSÃO
        comissao_regra = aplicar_regras_comissao(row)
        
        if comissao_regra is not None:
            return -comissao_regra if is_devolucao else comissao_regra
        
        # 3. TERCEIRO: OFERTAS (VOG e CB)
        comissao_oferta = buscar_oferta_vog(row, ofertas_off, ofertas_cb)
        
        if comissao_oferta is not None:
            return -comissao_oferta if is_devolucao else comissao_oferta
        
        # 4. QUARTO: VALOR PADRÃO (3%)
        return -0.03 if is_devolucao else 0.03
        
    except Exception as e:
        print(f"❌ ERRO no cálculo P.Com final: {e}, usando fallback 0.03")
        return -0.03 if is_devolucao else 0.03

# APLICAR O CÁLCULO FINAL DO P.COM
base_df['Comissão Kg'] = base_df.apply(calcular_comissao_kg_simplificada, axis=1)
base_df['P. Com'] = base_df.apply(calcular_p_com_final, axis=1)

# Verificar se algum P.Com ficou vazio
p_com_vazios = base_df['P. Com'].isna().sum()
if p_com_vazios > 0:
    print(f"⚠️  Existem {p_com_vazios} registros com P.Com vazio - aplicando fallback...")
    base_df['P. Com'] = base_df.apply(
        lambda row: -0.03 if str(row['CF']).startswith('DEV') else 0.03 
        if pd.isna(row['P. Com']) else row['P. Com'], 
        axis=1
    )

# Continuar com os cálculos das outras colunas...
base_df['Comissão Real'] = base_df.apply(
    lambda row: row['Comissão Kg'] * row['QTDE AJUSTADA'] if row['Preço Venda'] > 0 
    else -(row['Comissão Kg'] * row['QTDE AJUSTADA']), axis=1
)

base_df['Coleta Esc'] = base_df['Fat. Bruto'] * base_df['Escritório']
base_df['Frete Real'] = base_df['QTDE REAL2'] * base_df['Frete']

base_df['comissão'] = base_df.apply(
    lambda row: (row['P. Com']*row['Preço Venda']) / row['Preço Venda'] if row['Preço Venda'] > 0
    else -(row['P. Com']*row['Preço Venda']) / row['Preço Venda'] if row['Preço Venda'] < 0
    else 0, axis=1
)

base_df['Escr.'] = base_df.apply(
    lambda row: row['Coleta Esc'] / row['Fat. Bruto'] if row['Fat. Bruto'] != 0
    else 0, axis=1
)

base_df['frete'] = base_df.apply(
    lambda row: row['Frete Real'] / row['Fat. Bruto'] if row['Fat. Bruto'] != 0
    else 0, axis=1
)

base_df['TP'] = base_df.apply(
    lambda row: "BIG BACON" if str(row['CODPRODUTO']) == "700"
    else "REFFINATO" if row['GRUPO PRODUTO'] in ["SALGADOS SUINOS A GRANEL", "SALGADOS SUINOS EMBALADOS"]
    else "MIX", axis=1
)

base_df['CANC'] = base_df['NF-E'].apply(lambda x: "SIM" if x in notas_canceladas else "")

base_df['Armazenagem'] = base_df.apply(
    lambda row: (row['Fat. Bruto'] * row['P. Com']) / row['QTDE AJUSTADA'] if row['QTDE AJUSTADA'] != 0
    else 0, axis=1
)

def buscar_comissao_regra(row):
    try:
        if pd.notna(row['OS']) and pd.notna(row['NF-E']) and pd.notna(row['CODPRODUTO']):
            key = (int(row['OS']), int(row['NF-E']), int(row['CODPRODUTO']))
            return comissao_regra_dict.get(key, 0)
        else:
            return 0
    except:
        return 0

base_df['Comissão por Regra'] = base_df.apply(buscar_comissao_regra, axis=1)
base_df['Coluna2'] = base_df['Comissão por Regra'] == base_df['Comissão Kg']
base_df['FRETE - LUC/PREJ'] = base_df['QTDE AJUSTADA'] * base_df['Frete']

def buscar_desc_fec(row):
    try:
        pk = str(row['OS']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
        return fechamento_pk_dict.get(pk, {}).get('DESCONTO_VERIFICADO', np.nan)
    except:
        return np.nan

base_df['DESC FEC'] = base_df.apply(buscar_desc_fec, axis=1)

def buscar_esc_fec(row):
    try:
        pk = str(row['OS']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
        esc_fec_value = fechamento_pk_dict.get(pk, {}).get('ESCRITORIO', np.nan)
        if pd.isna(esc_fec_value):
            return np.nan
        esc_fec_value = float(esc_fec_value)
        # MODIFICAÇÃO: Transformar 4% e 0% em 3.5%
        if abs(esc_fec_value - 4.0) < 0.001 or abs(esc_fec_value - 0.0) < 0.001:
            esc_fec_value = 3.5
        return esc_fec_value / 100
    except:
        return np.nan

base_df['ESC FEC'] = base_df.apply(buscar_esc_fec, axis=1)

def buscar_icms_fec(row):
    try:
        pk = str(row['OS']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
        return fechamento_pk_dict.get(pk, {}).get('VLR ICMS', np.nan)
    except:
        return np.nan

base_df['ICMS FEC'] = base_df.apply(buscar_icms_fec, axis=1)

def buscar_prc_vend_fev(row):
    try:
        pk = str(row['OS']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
        return fechamento_pk_dict.get(pk, {}).get('PRECO VENDA', np.nan)
    except:
        return np.nan

base_df['PRC VEND FEV'] = base_df.apply(buscar_prc_vend_fev, axis=1)

base_df['DESC'] = base_df.apply(
    lambda row: (row['DESC FEC'] / 100) == row['Desc Perc'] if pd.notna(row['DESC FEC']) else False, axis=1
)

base_df['ESC'] = base_df.apply(
    lambda row: (row['ESC FEC'] / 100) == row['Escritório'] if pd.notna(row['ESC FEC']) else False, axis=1
)

base_df['ICMS'] = base_df.apply(
    lambda row: row['ICMS FEC'] == row['VL ICMS'] if pd.notna(row['ICMS FEC']) else False, axis=1
)

base_df['PRC VEND'] = base_df.apply(
    lambda row: row['PRC VEND FEV'] == row['Preço Venda'] if pd.notna(row['PRC VEND FEV']) else False, axis=1
)

def buscar_descricao_1(row):
    try:
        nf_e = int(row['NF-E']) if pd.notna(row['NF-E']) else None
        if nf_e is not None and nf_e in fechamento_nf_dict:
            return fechamento_nf_dict[nf_e]
        else:
            return ""
    except:
        return ""

base_df['DESCRIÇÃO_1'] = base_df.apply(buscar_descricao_1, axis=1)

base_df['MOV ENC'] = base_df.apply(
    lambda row: "ENCONTRADO" if any([str(row['OS']) == str(venda[0]) and str(row['NF-E']) == str(venda[1]) for venda in vendas_var])
    else "NÃO ENCONTRADO", axis=1
)

base_df['CUST + IMP'] = base_df['Custo real'] * base_df['QTDE AJUSTADA']
base_df['CUST PROD'] = base_df['QTDE AJUSTADA'] * base_df['Produção']
base_df['COM BRUTA'] = base_df['QTDE AJUSTADA'] * base_df['P. Com'] * base_df['Preço Venda']
base_df['Coluna1'] = (round(base_df['COM BRUTA'], 2) == round(base_df['Comissão Real'], 2))

base_df['Custo divergente'] = base_df.apply(
    lambda row: "CORRETO" if (row['QTDE'] > 0 and row['CUSTO EM SISTEMA'] == row['CUSTO']) else "DIVERGENTE", axis=1
)

def aplicar_regras_dev(row):
    if str(row['CF']).strip() == "DEV":
        colunas_zero = [
            'QTDE', 'CUSTO EM SISTEMA', 'CUSTO', 'Custo real', 'Frete', 
            'Produção', 'Escritório', 'Aniversário', 'Desc. Valor', 'Margem'
        ]
        
        colunas_negativas = [
            'P. Com', 'VL ICMS', 'Preço Venda', 'Fat Liquido', 
            'Fat. Bruto', 'Lucro / Prej.'
        ]
        
        for coluna in colunas_zero:
            if coluna in row.index:
                row[coluna] = 0
        
        for coluna in colunas_negativas:
            if coluna in row.index and row[coluna] != 0:
                if row[coluna] > 0:
                    row[coluna] = -row[coluna]
    
    return row

base_df = base_df.apply(aplicar_regras_dev, axis=1)

# Recalcular Lucro / Prej. e Margem após aplicar as regras
base_df['Lucro / Prej.'] = base_df.apply(
    lambda row: row['Fat. Bruto'] - (
        row['QTDE AJUSTADA'] * row['Custo real'] + 
        (row['Val Pis'] + row['VLRCOFINS'] + row['IRPJ'] +  row['CSLL'] + row['VL ICMS']) + 
        row['Desc. Valor'] + 
        (row['QTDE AJUSTADA'] * row['Frete']) + 
        (row['QTDE AJUSTADA'] * row['Produção']) + 
        (row['P. Com'] * row['Preço Venda'] * row['QTDE AJUSTADA']) + 
        row['Aniversário']
    ) - (row['Fat. Bruto'] * row['Escritório']), 
    axis=1
)

base_df['Margem'] = base_df.apply(
    lambda row: (row['Lucro / Prej.'] / row['Fat Liquido']) if row['Fat Liquido'] != 0 else 0, axis=1
)

base_df['Margem'] = base_df.apply(
    lambda row: 0 if str(row['CF']).strip() == "DEV" else row['Margem'], axis=1
)

base_df['INCL.'] = ""

def buscar_descricao_2(row):
    try:
        pk = str(row['OS']) + "_" + str(row['NF-E']) + "_" + str(row['CODPRODUTO'])
        return fechamento_pk_dict.get(pk, {}).get('MOV_V2', "")
    except:
        return ""

base_df['DESCRIÇÃO_2'] = base_df.apply(buscar_descricao_2, axis=1)

# Reordenar colunas
colunas_ordenadas = [
    'CF', 'RAZAO', 'FANTASIA', 'GRUPO', 'OS', 'NF-E', 'CF_NF', 'DATA', 'VENDEDOR',
    'CODPRODUTO', 'GRUPO PRODUTO', 'DESCRICAO', 'QTDE', 'QTDE REAL', 'CUSTO EM SISTEMA',
    'QTDE AJUSTADA', 'QTDE REAL2', 'CUSTO', 'Custo real', 'Custo total', 'Frete', 'Produção',
    'Escritório', 'Comissão Kg', 'P. Com', 'Aniversário', 'Val Pis', 'VLRCOFINS',
    'IRPJ', 'CSLL', 'VL ICMS', 'Aliq Icms', 'Desc Perc', 'Desc. Valor', 'Preço Venda',
    'Fat Liquido', 'Fat. Bruto', 'Lucro / Prej.', 'Margem', 'Quinzena', 'Comissão Real',
    'Coleta Esc', 'Frete Real', 'INCL.', 'comissão', 'Escr.', 'frete', 'Custo divergente',
    'TP', 'CANC', 'Armazenagem', 'Comissão por Regra', 'PK', 'Coluna2', 'FRETE - LUC/PREJ',
    'CUST + IMP', 'CUST PROD', 'COM BRUTA', 'DESC FEC', 'ESC FEC', 'ICMS FEC', 'PRC VEND FEV',
    'DESC', 'ESC', 'ICMS', 'PRC VEND', 'Coluna1', 'DESCRIÇÃO_1', 'DESCRIÇÃO_2'
]

colunas_existentes = [col for col in colunas_ordenadas if col in base_df.columns]
base_df = base_df[colunas_existentes]
base_df = base_df.fillna("")

# Salvar arquivos
print("💾 Salvando arquivos...")
output_path = f"C:\\Users\\win11\\Downloads\\Margem_{data_nome_arquivo}.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    base_df.to_excel(writer, sheet_name='Base (3,5%)', index=False)
    custos_produtos.to_excel(writer, sheet_name='PRECOS', index=False)
    cancelados.to_excel(writer, sheet_name='Base_cancelamento', index=False)
    devolucoes.to_excel(writer, sheet_name='Base_movimentacao', index=False)
    fechamento.to_excel(writer, sheet_name='Base_fechamento', index=False)
    
    if not lourencini.empty:
        lourencini.to_excel(writer, sheet_name='LOURENCINI', index=False)
    
    workbook = writer.book
    font_size = 10
    
    colunas_para_centralizar = [
        'CF', 'OS', 'NF-E', 'CF_NF', 'DATA', 'CODPRODUTO', 'QTDE', 'QTDE REAL', 
        'CUSTO EM SISTEMA', 'QTDE AJUSTADA', 'QTDE REAL2', 'Escritório', 'P. Com', 
        'Desc Perc', 'Margem', 'Quinzena', 'INCL.', 'comissão', 'Escr.', 'frete', 
        'Custo divergente', 'TP', 'PK', 'Coluna2', 'FRETE - LUC/PREJ', 'CUST PROD', 
        'COM BRUTA', 'DESC FEC', 'ESC FEC', 'ICMS FEC', 'PRC VEND FEV', 'DESC', 
        'ESC', 'ICMS', 'PRC VEND', 'Coluna1', 'DESCRIÇÃO_1', 'DESCRIÇÃO_2'
    ]
    
    colunas_monetarias = [
        'CUSTO', 'Custo real', 'Custo total', 'Frete', 'Produção', 'Comissão Kg', 'Aniversário',
        'VL ICMS', 'Desc. Valor', 'Preço Venda', 'Fat Liquido', 'Fat. Bruto',
        'Lucro / Prej.', 'Comissão Real', 'Coleta Esc', 'Frete Real',
        'Armazenagem', 'Comissão por Regra', 'CUST + IMP'
    ]
    
    colunas_porcentagem = [
        'Escritório', 'P. Com', 'Desc Perc', 'Margem', 'comissão', 'Escr.', 'frete',
        'DESC FEC', 'ESC FEC'
    ]
    
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        if sheet_name == 'Base (3,5%)':
            col_indices_monetarias = {}
            col_indices_porcentagem = {}
            col_indices_centralizar = {}
            
            for col_num in range(1, worksheet.max_column + 1):
                col_name = worksheet.cell(row=1, column=col_num).value
                if col_name in colunas_monetarias:
                    col_indices_monetarias[col_num] = col_name
                if col_name in colunas_porcentagem:
                    col_indices_porcentagem[col_num] = col_name
                if col_name in colunas_para_centralizar:
                    col_indices_centralizar[col_num] = col_name
        
        if sheet_name == 'Base (3,5%)':
            for col_num in col_indices_centralizar:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                
                for row_num in range(1, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row_num}']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if sheet_name == 'Base (3,5%)':
            for col_num in col_indices_monetarias:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row_num}']
                    if cell.value is not None:
                        try:
                            float(cell.value)
                            cell.number_format = '"R$"* #,##0.00;[Red]"R$"* -#,##0.00;"R$"* -'
                            if col_num in col_indices_centralizar:
                                cell.alignment = Alignment(horizontal='center', vertical='center', 
                                                         number_format='"R$"* #,##0.00;[Red]"R$"* -#,##0.00;"R$"* -')
                        except (ValueError, TypeError):
                            pass
        
        if sheet_name == 'Base (3,5%)':
            for col_num in col_indices_porcentagem:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row_num}']
                    if cell.value is not None:
                        try:
                            float(cell.value)
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='center', vertical='center', 
                                                     number_format='0.00%')
                        except (ValueError, TypeError):
                            pass
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font(size=font_size)
                if cell.row == 1 and cell.column in col_indices_porcentagem:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if any(char.isdigit() for char in str(cell.value)) and not any(char.isalpha() for char in str(cell.value)):
                            max_length = max(min(cell_length, 12), max_length)
                        else:
                            max_length = max(min(cell_length, 25), max_length)
                except:
                    pass
                
            adjusted_width = min(max_length + 2, 30)
            adjusted_width = max(adjusted_width, 10)
            worksheet.column_dimensions[column_letter].width = adjusted_width

# Salvar JSON
def default_serializer(obj):
    if isinstance(obj, (np.integer, int)):
        return int(obj)
    elif isinstance(obj, (np.floating, float)):
        return float(obj)
    elif isinstance(obj, (np.ndarray, pd.Series)):
        return obj.tolist()
    elif isinstance(obj, pd.Timestamp):
        return obj.isoformat() if not pd.isna(obj) else ""
    elif isinstance(obj, date):
        return obj.isoformat() if obj is not None else ""
    elif pd.isna(obj):
        return None
    elif obj in [np.inf, -np.inf]:
        return None
    elif callable(obj):
        return ""
    elif hasattr(obj, '__call__'):
        return ""
    raise TypeError(f"Type {type(obj)} not serializable")

json_path = f"C:\\Users\\win11\\Downloads\\Margem_{data_nome_arquivo}.json"
base_df_clean = base_df.copy()

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(base_df_clean.to_dict(orient='records'), f, ensure_ascii=False, indent=4, default=default_serializer)

print("✅ PROCESSAMENTO CONCLUÍDO!")
print(f"📄 Arquivo Excel: {output_path}")
print(f"📄 Arquivo JSON: {json_path}")
print(f"📊 Total de registros processados: {len(base_df)}")